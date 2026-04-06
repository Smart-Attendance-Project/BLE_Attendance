import io
import secrets
from datetime import datetime, date
import re

from fastapi import Depends, FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from sqlalchemy import inspect, text, and_, desc, func, or_, select
from sqlalchemy.orm import Session

from app.auth import (
    authenticate_user, create_access_token, get_current_user,
    hash_password, require_role, require_admin, require_super_admin,
)
from app.database import Base, engine, get_db
from app.models import (
    Attendance, Batch, Branch, Detection, Division, DivisionStudent,
    ScheduleSlot, Semester, Session as LectureSession, Subject,
    TeacherAssignment, User, UserRole,
)
from app.schemas import (
    AssignmentIn, AssignmentOut, AttendanceFinalizeIn, AttendanceOut,
    AttendanceOverrideIn, AttendanceStudentSummary, BatchDetectionIn,
    BatchIn, BatchOut, BranchIn, BranchOut, CreateAdminIn, CreateStudentIn,
    CreateTeacherIn, DetectionIn, DivisionIn, DivisionOut, DivisionStudentIn,
    LoginRequest, RegisterRequest, ScheduleSlotIn, ScheduleSlotOut,
    SemesterIn, SemesterOut, SessionAttendanceSummary, SessionCreateRequest,
    SessionOut, SubjectIn, SubjectOut, TokenResponse, UserOut,
)
from app.services import compute_presence_ratio, upsert_attendance

app = FastAPI(title="BLE Attendance API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

STUDENT_ID_PATTERN = re.compile(r"^(?:\d{2}[A-Z]{2,5}\d{3}|D\d{2}[A-Z]{2,5}\d{3})$")


def _run_migrations(engine):
    inspector = inspect(engine)
    tables = inspector.get_table_names()

    with engine.begin() as conn:
        conn.execute(text("ALTER TYPE userrole ADD VALUE IF NOT EXISTS 'admin'"))

        if "sessions" in tables:
            cols = {c["name"] for c in inspector.get_columns("sessions")}
            if "finalization_open" not in cols:
                conn.execute(text("ALTER TABLE sessions ADD COLUMN finalization_open BOOLEAN DEFAULT FALSE"))
            if "assignment_id" not in cols:
                conn.execute(text("ALTER TABLE sessions ADD COLUMN assignment_id INTEGER REFERENCES teacher_assignments(id)"))
            if "attendance_locked" not in cols:
                conn.execute(text("ALTER TABLE sessions ADD COLUMN attendance_locked BOOLEAN DEFAULT FALSE"))
        if "users" in tables:
            cols = {c["name"] for c in inspector.get_columns("users")}
            if "admin_id" not in cols:
                conn.execute(text("ALTER TABLE users ADD COLUMN admin_id VARCHAR(16) UNIQUE"))
            if "is_super_admin" not in cols:
                conn.execute(text("ALTER TABLE users ADD COLUMN is_super_admin BOOLEAN DEFAULT FALSE"))


@app.on_event("startup")
def on_startup() -> None:
    Base.metadata.create_all(bind=engine)
    _run_migrations(engine)


@app.get("/health")
def health_check():
    return {"status": "ok"}


# ── Auth ──────────────────────────────────────────────────────────────────────

@app.post("/auth/register", response_model=UserOut)
def register_user(payload: RegisterRequest, db: Session = Depends(get_db)):
    if payload.role == UserRole.student:
        if not STUDENT_ID_PATTERN.match(payload.identifier):
            raise HTTPException(status_code=422, detail="Invalid student ID format")
        if db.scalar(select(User).where(User.student_id == payload.identifier)):
            raise HTTPException(status_code=409, detail="Student ID already exists")
        user = User(full_name=payload.full_name, role=payload.role,
                    student_id=payload.identifier, password_hash=hash_password(payload.password))
    elif payload.role == UserRole.teacher:
        if db.scalar(select(User).where(User.teacher_id == payload.identifier)):
            raise HTTPException(status_code=409, detail="Teacher ID already exists")
        user = User(full_name=payload.full_name, role=payload.role,
                    teacher_id=payload.identifier, password_hash=hash_password(payload.password))
    else:
        raise HTTPException(status_code=400, detail="Use admin endpoints to create admin accounts")
    db.add(user)
    db.commit()
    db.refresh(user)
    return UserOut(id=user.id, full_name=user.full_name, role=user.role,
                   student_id=user.student_id, teacher_id=user.teacher_id)


@app.post("/auth/login", response_model=TokenResponse)
def login(payload: LoginRequest, db: Session = Depends(get_db)):
    user = authenticate_user(db, payload.identifier, payload.password)
    if user is None:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    token = create_access_token(subject=user.id)
    return TokenResponse(access_token=token, role=user.role,
                         full_name=user.full_name, user_id=user.id)


@app.get("/auth/me", response_model=UserOut)
def me(user: User = Depends(get_current_user)):
    return UserOut(id=user.id, full_name=user.full_name, role=user.role,
                   student_id=user.student_id, teacher_id=user.teacher_id,
                   admin_id=user.admin_id, is_super_admin=user.is_super_admin)


# ── BLE Sessions ──────────────────────────────────────────────────────────────

@app.post("/sessions", response_model=SessionOut)
def start_session(payload: SessionCreateRequest, db: Session = Depends(get_db),
                  teacher: User = Depends(require_role(UserRole.teacher))):
    active = db.scalar(select(LectureSession).where(
        and_(LectureSession.teacher_user_id == teacher.id, LectureSession.is_active.is_(True))
    ))
    if active:
        raise HTTPException(status_code=409, detail="A session is already active")
    session = LectureSession(subject=payload.subject, teacher_user_id=teacher.id,
                             token=secrets.token_hex(8), assignment_id=payload.assignment_id)
    db.add(session)
    db.commit()
    db.refresh(session)
    return _session_out(session)


@app.post("/sessions/{session_id}/end", response_model=SessionOut)
def end_session(session_id: str, db: Session = Depends(get_db),
                teacher: User = Depends(require_role(UserRole.teacher))):
    session = _get_own_session(db, session_id, teacher.id)
    session.is_active = False
    session.finalization_open = False
    session.ends_at = datetime.utcnow()
    db.commit()
    db.refresh(session)
    return _session_out(session)


@app.get("/sessions/active", response_model=SessionOut)
def get_active_session(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    session = db.scalar(select(LectureSession).where(LectureSession.is_active.is_(True))
                        .order_by(desc(LectureSession.starts_at)))
    if session is None:
        raise HTTPException(status_code=404, detail="No active session")
    return _session_out(session)


@app.post("/teacher/sessions/{session_id}/open-finalization", response_model=SessionOut)
def open_finalization(session_id: str, db: Session = Depends(get_db),
                      teacher: User = Depends(require_role(UserRole.teacher))):
    session = _get_own_session(db, session_id, teacher.id)
    if not session.is_active:
        raise HTTPException(status_code=409, detail="Session already ended")
    session.finalization_open = True
    db.commit()
    db.refresh(session)
    return _session_out(session)


# ── Detections ────────────────────────────────────────────────────────────────

@app.post("/detections")
def submit_detection(payload: DetectionIn, db: Session = Depends(get_db),
                     student: User = Depends(require_role(UserRole.student))):
    session = db.get(LectureSession, payload.session_id)
    if session is None or not session.is_active:
        raise HTTPException(status_code=404, detail="Active session not found")
    db.add(Detection(session_id=payload.session_id, student_user_id=student.id,
                     rssi=payload.rssi, proximity_ok=payload.proximity_ok))
    db.commit()
    return {"message": "Detection recorded"}


@app.post("/detections/batch")
def submit_detection_batch(payload: BatchDetectionIn, db: Session = Depends(get_db),
                           teacher: User = Depends(require_role(UserRole.teacher))):
    recorded = 0
    for item in payload.detections:
        session = db.get(LectureSession, item.session_id)
        if not session or not session.is_active or session.teacher_user_id != teacher.id:
            continue
        student = db.scalar(select(User).where(User.student_id == item.student_identifier))
        if not student:
            continue
        db.add(Detection(session_id=item.session_id, student_user_id=student.id,
                         rssi=item.rssi, proximity_ok=item.proximity_ok))
        recorded += 1
    db.commit()
    return {"message": f"{recorded} detections recorded"}


# ── Attendance finalize (student) ─────────────────────────────────────────────

@app.post("/attendance/finalize", response_model=AttendanceOut)
def finalize_attendance(payload: AttendanceFinalizeIn, db: Session = Depends(get_db),
                        student: User = Depends(require_role(UserRole.student))):
    session = db.get(LectureSession, payload.session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    if not session.finalization_open:
        raise HTTPException(status_code=409, detail="Finalization not open")
    att = upsert_attendance(db=db, session_id=payload.session_id,
                            student_user_id=student.id, biometric_verified=payload.biometric_verified)
    return AttendanceOut(session_id=att.session_id, student_user_id=att.student_user_id,
                         presence_ratio=att.presence_ratio, is_present=att.is_present,
                         biometric_verified=att.biometric_verified)


# ── Teacher: attendance summary + override + lock ─────────────────────────────

@app.get("/teacher/sessions/{session_id}/attendance-summary", response_model=SessionAttendanceSummary)
def get_attendance_summary(session_id: str, db: Session = Depends(get_db),
                           teacher: User = Depends(require_role(UserRole.teacher))):
    session = _get_own_session(db, session_id, teacher.id)
    return _build_summary(db, session)


@app.post("/teacher/sessions/{session_id}/attendance/override", response_model=AttendanceOut)
def override_attendance(session_id: str, payload: AttendanceOverrideIn,
                        db: Session = Depends(get_db),
                        teacher: User = Depends(require_role(UserRole.teacher))):
    session = _get_own_session(db, session_id, teacher.id)
    if session.attendance_locked:
        raise HTTPException(status_code=409, detail="Attendance is locked")
    student = db.get(User, payload.student_user_id)
    if not student or student.role != UserRole.student:
        raise HTTPException(status_code=404, detail="Student not found")
    att = db.scalar(select(Attendance).where(
        and_(Attendance.session_id == session_id, Attendance.student_user_id == payload.student_user_id)
    ))
    if att is None:
        att = Attendance(session_id=session_id, student_user_id=payload.student_user_id,
                         presence_ratio=0.0, biometric_verified=False)
        db.add(att)
    att.is_present = payload.is_present
    att.overridden_by_teacher = True
    att.override_reason = payload.reason
    db.commit()
    db.refresh(att)
    return AttendanceOut(session_id=att.session_id, student_user_id=att.student_user_id,
                         presence_ratio=att.presence_ratio, is_present=att.is_present,
                         biometric_verified=att.biometric_verified)


@app.post("/teacher/sessions/{session_id}/lock")
def lock_attendance(session_id: str, db: Session = Depends(get_db),
                    teacher: User = Depends(require_role(UserRole.teacher))):
    session = _get_own_session(db, session_id, teacher.id)
    session.attendance_locked = True
    db.commit()
    return {"message": "Attendance locked"}


@app.get("/teacher/sessions/{session_id}/detections")
def get_session_detections(session_id: str, db: Session = Depends(get_db),
                           teacher: User = Depends(require_role(UserRole.teacher))):
    session = _get_own_session(db, session_id, teacher.id)
    rows = db.execute(
        select(Detection, User).join(User, Detection.student_user_id == User.id)
        .where(Detection.session_id == session_id).order_by(desc(Detection.detected_at))
    ).all()
    return [{"student_id": u.student_id, "student_name": u.full_name,
             "rssi": d.rssi, "proximity_ok": d.proximity_ok,
             "detected_at": d.detected_at.isoformat()} for d, u in rows]


# ── Teacher: schedule & assignments ──────────────────────────────────────────

@app.get("/teacher/me/schedule/today")
def teacher_today_schedule(db: Session = Depends(get_db),
                           teacher: User = Depends(require_role(UserRole.teacher))):
    today = datetime.utcnow().weekday()  # 0=Mon
    sem = db.scalar(select(Semester).where(Semester.is_active.is_(True)))
    if not sem:
        return []
    slots = db.execute(
        select(ScheduleSlot).where(
            and_(ScheduleSlot.day_of_week == today, ScheduleSlot.semester_id == sem.id)
        ).order_by(ScheduleSlot.time_start)
    ).scalars().all()
    result = []
    for slot in slots:
        a = slot.assignment
        if a.teacher_user_id != teacher.id:
            continue
        result.append(_slot_out(slot))
    return result


@app.get("/teacher/me/assignments")
def teacher_assignments(db: Session = Depends(get_db),
                        teacher: User = Depends(require_role(UserRole.teacher))):
    rows = db.execute(
        select(TeacherAssignment).where(TeacherAssignment.teacher_user_id == teacher.id)
    ).scalars().all()
    return [_assignment_out(a) for a in rows]


@app.get("/teacher/me/sessions")
def teacher_sessions(
    from_date: str = Query(None),
    to_date: str = Query(None),
    assignment_id: int = Query(None),
    db: Session = Depends(get_db),
    teacher: User = Depends(require_role(UserRole.teacher)),
):
    q = select(LectureSession).where(LectureSession.teacher_user_id == teacher.id)
    if assignment_id:
        q = q.where(LectureSession.assignment_id == assignment_id)
    if from_date:
        q = q.where(LectureSession.starts_at >= datetime.fromisoformat(from_date))
    if to_date:
        q = q.where(LectureSession.starts_at <= datetime.fromisoformat(to_date + "T23:59:59"))
    sessions = db.execute(q.order_by(desc(LectureSession.starts_at))).scalars().all()
    return [_session_out(s) for s in sessions]


# ── Teacher: Excel export ─────────────────────────────────────────────────────

@app.get("/teacher/attendance/export")
def export_attendance_range(
    assignment_id: int = Query(...),
    from_date: str = Query(...),
    to_date: str = Query(...),
    db: Session = Depends(get_db),
    teacher: User = Depends(require_role(UserRole.teacher)),
):
    from openpyxl import Workbook

    assignment = db.get(TeacherAssignment, assignment_id)
    if not assignment or assignment.teacher_user_id != teacher.id:
        raise HTTPException(status_code=404, detail="Assignment not found")

    from_dt = datetime.fromisoformat(from_date)
    to_dt = datetime.fromisoformat(to_date + "T23:59:59")

    sessions = db.execute(
        select(LectureSession).where(
            and_(
                LectureSession.assignment_id == assignment_id,
                LectureSession.starts_at >= from_dt,
                LectureSession.starts_at <= to_dt,
                LectureSession.is_active.is_(False),
            )
        ).order_by(LectureSession.starts_at)
    ).scalars().all()

    # Get all students in this division (and batch if lab)
    q = select(User).join(DivisionStudent, DivisionStudent.student_user_id == User.id).where(
        DivisionStudent.division_id == assignment.division_id
    )
    if assignment.batch_id:
        q = q.where(DivisionStudent.batch_id == assignment.batch_id)
    students = db.execute(q.order_by(User.student_id)).scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # Header row: ID | date1 | date2 | ...
    ws.cell(row=1, column=1, value="ID")
    for col, s in enumerate(sessions, 2):
        ws.cell(row=1, column=col, value=s.starts_at.strftime("%d/%m/%Y"))

    # Data rows
    for row, student in enumerate(students, 2):
        ws.cell(row=row, column=1, value=student.student_id)
        for col, s in enumerate(sessions, 2):
            att = db.scalar(select(Attendance).where(
                and_(Attendance.session_id == s.id, Attendance.student_user_id == student.id)
            ))
            ws.cell(row=row, column=col, value="P" if (att and att.is_present) else "A")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    fname = f"attendance_{assignment.subject.code}_{from_date}_to_{to_date}.xlsx"
    return StreamingResponse(output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# ── Teacher: student management ───────────────────────────────────────────────

@app.post("/teacher/students", response_model=UserOut)
def teacher_add_student(payload: CreateStudentIn, db: Session = Depends(get_db),
                        teacher: User = Depends(require_role(UserRole.teacher))):
    if not STUDENT_ID_PATTERN.match(payload.student_id):
        raise HTTPException(status_code=422, detail="Invalid student ID format")
    if db.scalar(select(User).where(User.student_id == payload.student_id)):
        raise HTTPException(status_code=409, detail="Student ID already exists")
    student = User(full_name=payload.full_name, role=UserRole.student,
                   student_id=payload.student_id, password_hash=hash_password(payload.password))
    db.add(student)
    db.flush()
    if payload.division_id:
        db.add(DivisionStudent(student_user_id=student.id,
                               division_id=payload.division_id, batch_id=payload.batch_id))
    db.commit()
    db.refresh(student)
    return UserOut(id=student.id, full_name=student.full_name, role=student.role,
                   student_id=student.student_id)


@app.get("/teacher/divisions/{division_id}/students")
def get_division_students(division_id: int, db: Session = Depends(get_db),
                          _: User = Depends(require_role(UserRole.teacher))):
    rows = db.execute(
        select(User, DivisionStudent)
        .join(DivisionStudent, DivisionStudent.student_user_id == User.id)
        .where(DivisionStudent.division_id == division_id)
        .order_by(User.student_id)
    ).all()
    return [{"id": u.id, "student_id": u.student_id, "full_name": u.full_name,
             "batch_id": ds.batch_id} for u, ds in rows]


# ── Admin: teachers & admins ──────────────────────────────────────────────────

@app.post("/admin/teachers", response_model=UserOut)
def admin_create_teacher(payload: CreateTeacherIn, db: Session = Depends(get_db),
                         _: User = Depends(require_admin)):
    if db.scalar(select(User).where(User.teacher_id == payload.teacher_id)):
        raise HTTPException(status_code=409, detail="Teacher ID already exists")
    user = User(full_name=payload.full_name, role=UserRole.teacher,
                teacher_id=payload.teacher_id, password_hash=hash_password(payload.password))
    db.add(user)
    db.commit()
    db.refresh(user)
    return UserOut(id=user.id, full_name=user.full_name, role=user.role, teacher_id=user.teacher_id)


@app.get("/admin/teachers")
def admin_list_teachers(db: Session = Depends(get_db), _: User = Depends(require_admin)):
    users = db.execute(select(User).where(User.role == UserRole.teacher)).scalars().all()
    return [{"id": u.id, "full_name": u.full_name, "teacher_id": u.teacher_id} for u in users]


@app.post("/admin/admins", response_model=UserOut)
def admin_create_admin(payload: CreateAdminIn, db: Session = Depends(get_db),
                       current: User = Depends(require_admin)):
    if payload.is_super_admin and not current.is_super_admin:
        raise HTTPException(status_code=403, detail="Only super-admin can create super-admins")
    if db.scalar(select(User).where(User.admin_id == payload.admin_id)):
        raise HTTPException(status_code=409, detail="Admin ID already exists")
    user = User(full_name=payload.full_name, role=UserRole.admin,
                admin_id=payload.admin_id, password_hash=hash_password(payload.password),
                is_super_admin=payload.is_super_admin)
    db.add(user)
    db.commit()
    db.refresh(user)
    return UserOut(id=user.id, full_name=user.full_name, role=user.role,
                   admin_id=user.admin_id, is_super_admin=user.is_super_admin)


@app.get("/admin/admins")
def admin_list_admins(db: Session = Depends(get_db), _: User = Depends(require_admin)):
    users = db.execute(select(User).where(User.role == UserRole.admin)).scalars().all()
    return [{"id": u.id, "full_name": u.full_name, "admin_id": u.admin_id,
             "is_super_admin": u.is_super_admin} for u in users]


# ── Admin: Semesters ──────────────────────────────────────────────────────────

@app.post("/admin/semesters", response_model=SemesterOut)
def create_semester(payload: SemesterIn, db: Session = Depends(get_db),
                    _: User = Depends(require_admin)):
    if payload.is_active:
        db.execute(text("UPDATE semesters SET is_active = FALSE"))
    sem = Semester(name=payload.name, start_date=payload.start_date,
                   end_date=payload.end_date, is_active=payload.is_active)
    db.add(sem)
    db.commit()
    db.refresh(sem)
    return _sem_out(sem)


@app.get("/admin/semesters", response_model=list[SemesterOut])
def list_semesters(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    sems = db.execute(select(Semester).order_by(Semester.id.desc())).scalars().all()
    return [_sem_out(s) for s in sems]


@app.patch("/admin/semesters/{sem_id}/activate")
def activate_semester(sem_id: int, db: Session = Depends(get_db),
                      _: User = Depends(require_admin)):
    db.execute(text("UPDATE semesters SET is_active = FALSE"))
    sem = db.get(Semester, sem_id)
    if not sem:
        raise HTTPException(status_code=404, detail="Semester not found")
    sem.is_active = True
    db.commit()
    return _sem_out(sem)


# ── Admin: Branches ───────────────────────────────────────────────────────────

@app.post("/admin/branches", response_model=BranchOut)
def create_branch(payload: BranchIn, db: Session = Depends(get_db),
                  _: User = Depends(require_admin)):
    b = Branch(code=payload.code, name=payload.name)
    db.add(b)
    db.commit()
    db.refresh(b)
    return BranchOut(id=b.id, code=b.code, name=b.name)


@app.get("/admin/branches", response_model=list[BranchOut])
def list_branches(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    rows = db.execute(select(Branch).order_by(Branch.code)).scalars().all()
    return [BranchOut(id=b.id, code=b.code, name=b.name) for b in rows]


# ── Admin: Divisions ──────────────────────────────────────────────────────────

@app.post("/admin/divisions", response_model=DivisionOut)
def create_division(payload: DivisionIn, db: Session = Depends(get_db),
                    _: User = Depends(require_admin)):
    d = Division(branch_id=payload.branch_id, year=payload.year,
                 div_number=payload.div_number, label=payload.label)
    db.add(d)
    db.commit()
    db.refresh(d)
    return _div_out(d)


@app.get("/admin/divisions", response_model=list[DivisionOut])
def list_divisions(branch_id: int = Query(None), db: Session = Depends(get_db),
                   _: User = Depends(get_current_user)):
    q = select(Division)
    if branch_id:
        q = q.where(Division.branch_id == branch_id)
    rows = db.execute(q.order_by(Division.year, Division.div_number)).scalars().all()
    return [_div_out(d) for d in rows]


# ── Admin: Batches ────────────────────────────────────────────────────────────

@app.post("/admin/batches", response_model=BatchOut)
def create_batch(payload: BatchIn, db: Session = Depends(get_db),
                 _: User = Depends(require_admin)):
    b = Batch(division_id=payload.division_id, label=payload.label)
    db.add(b)
    db.commit()
    db.refresh(b)
    return BatchOut(id=b.id, division_id=b.division_id, label=b.label)


@app.get("/admin/batches", response_model=list[BatchOut])
def list_batches(division_id: int = Query(None), db: Session = Depends(get_db),
                 _: User = Depends(get_current_user)):
    q = select(Batch)
    if division_id:
        q = q.where(Batch.division_id == division_id)
    rows = db.execute(q.order_by(Batch.label)).scalars().all()
    return [BatchOut(id=b.id, division_id=b.division_id, label=b.label) for b in rows]


# ── Admin: Subjects ───────────────────────────────────────────────────────────

@app.post("/admin/subjects", response_model=SubjectOut)
def create_subject(payload: SubjectIn, db: Session = Depends(get_db),
                   _: User = Depends(require_admin)):
    s = Subject(code=payload.code, name=payload.name, subject_type=payload.subject_type)
    db.add(s)
    db.commit()
    db.refresh(s)
    return SubjectOut(id=s.id, code=s.code, name=s.name, subject_type=s.subject_type)


@app.get("/admin/subjects", response_model=list[SubjectOut])
def list_subjects(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    rows = db.execute(select(Subject).order_by(Subject.code)).scalars().all()
    return [SubjectOut(id=s.id, code=s.code, name=s.name, subject_type=s.subject_type) for s in rows]


# ── Admin: Assignments ────────────────────────────────────────────────────────

@app.post("/admin/assignments", response_model=AssignmentOut)
def create_assignment(payload: AssignmentIn, db: Session = Depends(get_db),
                      _: User = Depends(require_admin)):
    a = TeacherAssignment(teacher_user_id=payload.teacher_user_id,
                          subject_id=payload.subject_id, division_id=payload.division_id,
                          batch_id=payload.batch_id)
    db.add(a)
    db.commit()
    db.refresh(a)
    return _assignment_out(a)


@app.get("/admin/assignments", response_model=list[AssignmentOut])
def list_assignments(teacher_id: str = Query(None), division_id: int = Query(None),
                     db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    q = select(TeacherAssignment)
    if teacher_id:
        q = q.where(TeacherAssignment.teacher_user_id == teacher_id)
    if division_id:
        q = q.where(TeacherAssignment.division_id == division_id)
    rows = db.execute(q).scalars().all()
    return [_assignment_out(a) for a in rows]


@app.delete("/admin/assignments/{assignment_id}")
def delete_assignment(assignment_id: int, db: Session = Depends(get_db),
                      _: User = Depends(require_admin)):
    a = db.get(TeacherAssignment, assignment_id)
    if not a:
        raise HTTPException(status_code=404, detail="Not found")
    db.delete(a)
    db.commit()
    return {"message": "Deleted"}


# ── Admin: Schedule Slots ─────────────────────────────────────────────────────

@app.post("/admin/schedule-slots", response_model=ScheduleSlotOut)
def create_slot(payload: ScheduleSlotIn, db: Session = Depends(get_db),
                _: User = Depends(require_admin)):
    slot = ScheduleSlot(assignment_id=payload.assignment_id, semester_id=payload.semester_id,
                        day_of_week=payload.day_of_week, time_start=payload.time_start,
                        time_end=payload.time_end, room=payload.room)
    db.add(slot)
    db.commit()
    db.refresh(slot)
    return _slot_out(slot)


@app.get("/admin/schedule-slots", response_model=list[ScheduleSlotOut])
def list_slots(semester_id: int = Query(None), division_id: int = Query(None),
               teacher_id: str = Query(None), db: Session = Depends(get_db),
               _: User = Depends(get_current_user)):
    q = select(ScheduleSlot)
    if semester_id:
        q = q.where(ScheduleSlot.semester_id == semester_id)
    rows = db.execute(q).scalars().all()
    result = []
    for slot in rows:
        a = slot.assignment
        if division_id and a.division_id != division_id:
            continue
        if teacher_id and a.teacher_user_id != teacher_id:
            continue
        result.append(_slot_out(slot))
    return result


@app.put("/admin/schedule-slots/{slot_id}", response_model=ScheduleSlotOut)
def update_slot(slot_id: int, payload: ScheduleSlotIn, db: Session = Depends(get_db),
                _: User = Depends(require_admin)):
    slot = db.get(ScheduleSlot, slot_id)
    if not slot:
        raise HTTPException(status_code=404, detail="Not found")
    slot.assignment_id = payload.assignment_id
    slot.semester_id = payload.semester_id
    slot.day_of_week = payload.day_of_week
    slot.time_start = payload.time_start
    slot.time_end = payload.time_end
    slot.room = payload.room
    db.commit()
    db.refresh(slot)
    return _slot_out(slot)


@app.delete("/admin/schedule-slots/{slot_id}")
def delete_slot(slot_id: int, db: Session = Depends(get_db),
                _: User = Depends(require_admin)):
    slot = db.get(ScheduleSlot, slot_id)
    if not slot:
        raise HTTPException(status_code=404, detail="Not found")
    db.delete(slot)
    db.commit()
    return {"message": "Deleted"}


# ── Admin: Students ───────────────────────────────────────────────────────────

@app.post("/admin/students", response_model=UserOut)
def admin_create_student(payload: CreateStudentIn, db: Session = Depends(get_db),
                         _: User = Depends(require_admin)):
    if not STUDENT_ID_PATTERN.match(payload.student_id):
        raise HTTPException(status_code=422, detail="Invalid student ID format")
    if db.scalar(select(User).where(User.student_id == payload.student_id)):
        raise HTTPException(status_code=409, detail="Student ID already exists")
    student = User(full_name=payload.full_name, role=UserRole.student,
                   student_id=payload.student_id, password_hash=hash_password(payload.password))
    db.add(student)
    db.flush()
    if payload.division_id:
        db.add(DivisionStudent(student_user_id=student.id,
                               division_id=payload.division_id, batch_id=payload.batch_id))
    db.commit()
    db.refresh(student)
    return UserOut(id=student.id, full_name=student.full_name, role=student.role,
                   student_id=student.student_id)


@app.post("/admin/division-students")
def assign_student_to_division(payload: DivisionStudentIn, db: Session = Depends(get_db),
                                _: User = Depends(require_admin)):
    existing = db.scalar(select(DivisionStudent).where(
        and_(DivisionStudent.student_user_id == payload.student_user_id,
             DivisionStudent.division_id == payload.division_id)
    ))
    if existing:
        existing.batch_id = payload.batch_id
    else:
        db.add(DivisionStudent(student_user_id=payload.student_user_id,
                               division_id=payload.division_id, batch_id=payload.batch_id))
    db.commit()
    return {"message": "Assigned"}


# ── Admin: attendance override ────────────────────────────────────────────────

@app.post("/admin/sessions/{session_id}/attendance/override", response_model=AttendanceOut)
def admin_override_attendance(session_id: str, payload: AttendanceOverrideIn,
                               db: Session = Depends(get_db), _: User = Depends(require_admin)):
    session = db.get(LectureSession, session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    if session.attendance_locked:
        raise HTTPException(status_code=409, detail="Attendance is locked")
    att = db.scalar(select(Attendance).where(
        and_(Attendance.session_id == session_id, Attendance.student_user_id == payload.student_user_id)
    ))
    if att is None:
        att = Attendance(session_id=session_id, student_user_id=payload.student_user_id,
                         presence_ratio=0.0, biometric_verified=False)
        db.add(att)
    att.is_present = payload.is_present
    att.overridden_by_teacher = True
    att.override_reason = payload.reason
    db.commit()
    db.refresh(att)
    return AttendanceOut(session_id=att.session_id, student_user_id=att.student_user_id,
                         presence_ratio=att.presence_ratio, is_present=att.is_present,
                         biometric_verified=att.biometric_verified)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _get_own_session(db, session_id, teacher_id):
    session = db.get(LectureSession, session_id)
    if not session or session.teacher_user_id != teacher_id:
        raise HTTPException(status_code=404, detail="Session not found")
    return session


def _session_out(s: LectureSession) -> SessionOut:
    return SessionOut(id=s.id, subject=s.subject, token=s.token,
                      teacher_user_id=s.teacher_user_id, assignment_id=s.assignment_id,
                      starts_at=s.starts_at, ends_at=s.ends_at, is_active=s.is_active,
                      finalization_open=s.finalization_open,
                      attendance_locked=s.attendance_locked or False)


def _sem_out(s: Semester) -> SemesterOut:
    return SemesterOut(id=s.id, name=s.name,
                       start_date=str(s.start_date), end_date=str(s.end_date),
                       is_active=s.is_active)


def _div_out(d: Division) -> DivisionOut:
    return DivisionOut(id=d.id, branch_id=d.branch_id, year=d.year,
                       div_number=d.div_number, label=d.label,
                       branch_code=d.branch.code if d.branch else "")


def _assignment_out(a: TeacherAssignment) -> AssignmentOut:
    return AssignmentOut(
        id=a.id, teacher_user_id=a.teacher_user_id,
        teacher_name=a.teacher.full_name if a.teacher else "",
        subject_id=a.subject_id,
        subject_name=a.subject.name if a.subject else "",
        subject_code=a.subject.code if a.subject else "",
        division_id=a.division_id,
        division_label=a.division.label if a.division else "",
        batch_id=a.batch_id,
        batch_label=a.batch.label if a.batch else None,
    )


def _slot_out(slot: ScheduleSlot) -> ScheduleSlotOut:
    a = slot.assignment
    return ScheduleSlotOut(
        id=slot.id, assignment_id=slot.assignment_id, semester_id=slot.semester_id,
        day_of_week=slot.day_of_week, time_start=slot.time_start, time_end=slot.time_end,
        room=slot.room,
        subject_name=a.subject.name if a and a.subject else "",
        subject_code=a.subject.code if a and a.subject else "",
        division_label=a.division.label if a and a.division else "",
        batch_label=a.batch.label if a and a.batch else None,
        teacher_name=a.teacher.full_name if a and a.teacher else "",
    )


def _build_summary(db, session: LectureSession) -> SessionAttendanceSummary:
    rows = db.execute(
        select(User, Attendance, func.count(Detection.id).label("dc"))
        .select_from(User)
        .outerjoin(Detection, and_(Detection.student_user_id == User.id,
                                   Detection.session_id == session.id))
        .outerjoin(Attendance, and_(Attendance.student_user_id == User.id,
                                    Attendance.session_id == session.id))
        .where(User.role == UserRole.student)
        .having(or_(func.count(Detection.id) > 0, Attendance.id.is_not(None)))
        .group_by(User.id, Attendance.id)
        .order_by(User.student_id)
    ).all()

    records, present = [], 0
    for user, att, dc in rows:
        if att is None:
            ratio = compute_presence_ratio(db, session.id, user.id)
            is_present = ratio >= 0.6
            bio, overridden, reason = False, False, None
        else:
            ratio, is_present = att.presence_ratio, att.is_present
            bio, overridden, reason = att.biometric_verified, att.overridden_by_teacher, att.override_reason
        if is_present:
            present += 1
        records.append(AttendanceStudentSummary(
            student_user_id=user.id, student_id=user.student_id, student_name=user.full_name,
            detection_count=dc, presence_ratio=ratio, is_present=is_present,
            biometric_verified=bio, overridden_by_teacher=overridden, override_reason=reason,
        ))
    return SessionAttendanceSummary(session_id=session.id, subject=session.subject,
                                    starts_at=session.starts_at, ends_at=session.ends_at,
                                    total_students=len(records), present_students=present,
                                    records=records)
